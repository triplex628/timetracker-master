from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

from django.http import HttpResponse
from api.models import TaskModel, ItemModel

import os
import datetime

from django.utils import timezone

class Detail:
    def __init__(self, types_of_work, subtypes_of_work, category_of_item, subcategory_of_item, title, seria):
        self.types_of_work = types_of_work
        self.subtypes_of_work = subtypes_of_work
        self.category_of_item = category_of_item
        self.subcategory_of_item = subcategory_of_item
        self.title = title
        self.seria = seria

    def __str__(self):
        return f'{self.types_of_work}; {self.subtypes_of_work}; {self.category_of_item}; {self.subcategory_of_item}; ' \
               f'{self.title}; {self.seria}'


class EmployeeTaskInfo:
    def __init__(self):
        self.total_time = datetime.timedelta(milliseconds=0)
        self.employees = []


class ReportGenerator:
    def __init__(self):
        self.details = []
        self.filtered_items = ItemModel.objects.all()

    def get_required_verb(self, type_of_action: str) -> str:
        match type_of_action:
            case 'Производство':
                return 'произведен'
            case 'Тестирование':
                return 'протестирован'
            case 'Ремонт':
                return 'отремонтирован'
            case 'Отгрузка':
                return 'выгружен'
            case _:
                return ''
            
    def get_required_verb_task(self, type_of_action: str) -> str:
        match type_of_action:
            case 'Производство':
                return 'Произвести'
            case 'Тестирование':
                return 'Протестировать'
            case 'Ремонт':
                return 'Отремонтировать'
            case 'Отгрузка':
                return 'Выгрузить'
            case _:
                return ''

    def upload_items(self):
        excel_file = os.path.join('templates', 'Заполнение.xlsx')
        wb = load_workbook(excel_file)

        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]

            data = []
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=4):
                row_data = []
                for cell in row:
                    cell_data = cell.value
                    row_data.append(cell_data)
                data.append(row_data)

            cur_subtypes_of_item = None
            cur_category_of_item = None
            cur_subcategory_of_item = None
            for row in data:
                element_type = row[-1]
                match element_type:
                    case 'SUBTYPES_OF_WORK':
                        cur_subtypes_of_item = row[0].strip()
                        cur_category_of_item = None
                        cur_subcategory_of_item = None
                    case 'CATEGORY_OF_ITEM':
                        cur_category_of_item = row[1].strip()
                        cur_subcategory_of_item = None
                    case 'SUBCATEGORY_OF_ITEM':
                        cur_subcategory_of_item = row[1].strip()
                    case None:
                        self.details.append(
                            Detail(
                                sheetname, 
                                cur_subtypes_of_item, 
                                cur_category_of_item, 
                                cur_subcategory_of_item, 
                                row[1], 
                                row[2]
                            )
                        )

    def generate_filtered_items(self):
        for detail in self.details:
            item = ItemModel.objects.filter(
                types_of_work=detail.types_of_work, 
                subtypes_of_work=detail.subtypes_of_work,
                category_of_item=detail.category_of_item,
                subcategory_of_item=detail.subcategory_of_item,
                title=detail.title,
                seria = detail.seria,
            ).first()

            filtered_tasks = TaskModel.objects.filter(item=item)
            if filtered_tasks:
                self.filtered_items.append(item)

#  generate first sheet
    def generate_first_sheet(self, sheet, employee_tasks):
        sheet['B1'] = "Задача"
        sheet['C1'] = "Результат"
        sheet['D1'] = "Потраченное время"
        sheet['E1'] = "Дата"
        sheet['F1'] = "ID администартора"
        sheet['G1'] = "ID прибора"

        for index in range(6):
            column_letter = chr(ord('A') + index)
            sheet.column_dimensions[column_letter].auto_size = True

        tasks = {}
        for employee_task in employee_tasks:
            if not (employee_task.task.id in tasks):
                tasks[employee_task.task.id] = []
            tasks[employee_task.task.id].append(employee_task)

        row_index = 2
        for employee_task in employee_tasks:
            task = employee_task.task
            total_time = employee_task.total_time
            min_date = employee_task.end_time.replace(tzinfo=None)
            admin_id = task.admin_id
            item_id = task.manual_item_id
            
            sheet[f'A{row_index}'] = row_index - 1
            sheet[f'B{row_index}'] = f'{self.get_required_verb_task(task.get_type_of_task_display())} {task.item.title}'
            sheet[f'C{row_index}'] = f'{task.item.title} {self.get_required_verb(task.get_type_of_task_display())}' 
            sheet[f'D{row_index}'] = total_time
            sheet[f'E{row_index}'] = min_date.strftime('%d.%m.%Y')
            sheet[f'F{row_index}'] = admin_id
            sheet[f'G{row_index}'] = item_id
            row_index += 1

#   generate second sheet
    def generate_second_sheet(self, sheet, employee_tasks):
        sheet['B1'] = 'Изделие'
        sheet['C1'] = 'Тип работы'
        sheet['D1'] = 'Продолжительность'

        row_index = 2
        for index in range(4):
            column_letter = chr(ord('A') + index)
            sheet.column_dimensions[column_letter].auto_size = True
        
        for employee_task in employee_tasks:
            task = employee_task.task
            sheet[f'A{row_index}'] = row_index - 1
            sheet[f'B{row_index}'] = task.title
            sheet[f'C{row_index}'] = task.get_type_of_task_display()
            sheet[f'D{row_index}'] = employee_task.total_time 
            row_index += 1

    #   the third sheet
    def generate_third_sheet(self, sheet, employee_tasks):
        for index in range(8):
            column_letter = chr(ord('A') + index)
            sheet.column_dimensions[column_letter].auto_size = True

        sheet.merge_cells('E1:I1')
        sheet['E1'] = 'Тип работ'
        sheet['E1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['E1'].font = Font(size=14, bold=True)

        types_of_task = ['Производство', 'Ремонт', 'Тестирование', 'Отгрузка', 'Итого']
        for i in range(len(types_of_task)):
            sheet[f'{chr(ord("E") + i)}2'] = types_of_task[i]

        cur_types_of_work = None
        cur_subtypes_of_work = None
        cur_category_of_item = None
        cur_subcategory_of_item = None
    
        row_index = 1

        for detail in self.filtered_items:
            flag = False
            item = ItemModel.objects.filter(
                types_of_work=detail.types_of_work, 
                subtypes_of_work=detail.subtypes_of_work,
                category_of_item=detail.category_of_item,
                subcategory_of_item=detail.subcategory_of_item,
                title=detail.title,
                seria = detail.seria,
            ).first()

            filtered_tasks = TaskModel.objects.filter(item=item)

            for filtered_task in filtered_tasks:
                filtered_employee_tasks = employee_tasks.filter(task=filtered_task)
                
                if not filtered_employee_tasks:
                    flag = True
                    break
                    
                if item.subtypes_of_work != cur_subtypes_of_work:
                    cur_subtypes_of_work = item.subtypes_of_work
                    cur_category_of_item = None
                    cur_subcategory_of_item = None
                    
                    cell_index = f'A{row_index}'
                    sheet.merge_cells(f'{cell_index}:D{row_index}')
                    sheet[cell_index].alignment = Alignment(horizontal='center', vertical='center')
                    sheet[cell_index].font = Font(size=11, bold=True)
                    sheet[cell_index] = cur_subtypes_of_work
                    row_index += 1

                if item.category_of_item != cur_category_of_item:
                    cur_category_of_item = item.category_of_item
                    cur_subcategory_of_item = None

                    cell_index = f'B{row_index}'
                    sheet.merge_cells(f'{cell_index}:D{row_index}')
                    sheet[cell_index].alignment = Alignment(horizontal='center', vertical='center')
                    sheet[cell_index].font = Font(size=11, bold=True)
                    sheet[cell_index] = cur_category_of_item
                    row_index += 1

                if item.subcategory_of_item != cur_subcategory_of_item:
                    cur_subcategory_of_item = item.subcategory_of_item
                    
                    cell_index = f'B{row_index}'
                    sheet.merge_cells(f'{cell_index}:D{row_index}')
                    sheet[cell_index].alignment = Alignment(horizontal='center', vertical='center')
                    sheet[cell_index].font = Font(size=11, bold=True)
                    sheet[cell_index] = cur_subcategory_of_item
                    row_index += 1
                
                sheet[f'C{row_index}'].alignment = Alignment(horizontal='center', vertical='center')
                sheet[f'C{row_index}'].font = Font(size=11, bold=True)
                sheet[f'C{row_index}'] = item.title

                sheet[f'D{row_index}'].alignment = Alignment(horizontal='center', vertical='center')
                sheet[f'D{row_index}'].font = Font(size=11, bold=True)
                sheet[f'D{row_index}'] = item.seria
                
                employee_tasks_info = {
                    'Производство': EmployeeTaskInfo(),
                    'Ремонт': EmployeeTaskInfo(), 
                    'Тестирование': EmployeeTaskInfo(), 
                    'Отгрузка': EmployeeTaskInfo(),
                    'Прочее': EmployeeTaskInfo() 
                }

                for filtered_employee_task in filtered_employee_tasks:
                    employee_tasks_info[filtered_task.get_type_of_task_display()].total_time += filtered_employee_task.total_time 
                    employee_tasks_info[filtered_task.get_type_of_task_display()].employees.append(f'{filtered_employee_task.employee}')

                col_index = 0
                total_time = datetime.timedelta(milliseconds=0)
                for key, value in employee_tasks_info.items():
                    sheet[f'{chr(ord("E") + col_index)}{row_index}'] = value.total_time
                    total_time += value.total_time
                    sheet[f'{chr(ord("E") + col_index)}{row_index + 1}'] = '\n'.join(value.employees)
                    col_index += 1
                if not flag:
                    sheet[f'{chr(ord("E") + col_index)}{row_index}'] = total_time
                    row_index += 2

    def generate_report(self, employee_tasks):
        workbook = Workbook()

        self.generate_filtered_items()
        sheet1 = workbook.active
        sheet1.title = "Лист 1"
        self.generate_first_sheet(sheet1, employee_tasks)

        sheet2 = workbook.create_sheet(title="Лист 2")
        workbook.active = sheet2
        self.generate_second_sheet(sheet2, employee_tasks)

        sheet3 = workbook.create_sheet(title="Лист 3")
        workbook.active = sheet3
        self.generate_third_sheet(sheet3, employee_tasks)

        workbook.active = sheet1

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        response['Content-Disposition'] = 'attachment; filename=example.xlsx'
        workbook.save(response) 
        # workbook.save('Мой_файл.xlsx')       
        return response
